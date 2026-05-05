"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice (optional).
Falls back gracefully when LibreOffice is unavailable.
"""

import json
import os
import platform
import shutil
import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl is not installed. Run: pip install openpyxl"}))
    sys.exit(1)

MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_DIR_WINDOWS = os.path.join(os.environ.get("APPDATA", ""), "LibreOffice", "4", "user", "basic", "Standard")
MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def get_soffice_path():
    """Find LibreOffice soffice binary across platforms."""
    # Check PATH first
    soffice = shutil.which("soffice")
    if soffice:
        return soffice

    # Platform-specific default locations
    system = platform.system()
    candidates = []
    if system == "Darwin":
        candidates = [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        ]
    elif system == "Windows":
        program_files = os.environ.get("PROGRAMFILES", "C:\\Program Files")
        program_files_x86 = os.environ.get("PROGRAMFILES(X86)", "C:\\Program Files (x86)")
        candidates = [
            os.path.join(program_files, "LibreOffice", "program", "soffice.exe"),
            os.path.join(program_files_x86, "LibreOffice", "program", "soffice.exe"),
        ]
    else:  # Linux
        candidates = [
            "/usr/bin/soffice",
            "/usr/lib/libreoffice/program/soffice",
        ]

    for c in candidates:
        if os.path.isfile(c):
            return c

    return None


def get_soffice_env():
    """Return a clean environment for LibreOffice to avoid conflicts."""
    env = os.environ.copy()
    # Remove Python-related env vars that can confuse LibreOffice's internal Python
    for key in list(env.keys()):
        if key.startswith("PYTHON") or key == "VIRTUAL_ENV":
            del env[key]
    return env


def has_timeout_cmd():
    """Check for timeout command availability (gtimeout on macOS, timeout on Linux)."""
    system = platform.system()
    if system == "Darwin":
        try:
            subprocess.run(["gtimeout", "--version"], capture_output=True, timeout=2, check=False)
            return "gtimeout"
        except (FileNotFoundError, subprocess.TimeoutExpired):
            return None
    elif system == "Linux":
        try:
            subprocess.run(["timeout", "--version"], capture_output=True, timeout=2, check=False)
            return "timeout"
        except (FileNotFoundError, subprocess.TimeoutExpired):
            return None
    # Windows: no standard timeout wrapper for subprocess
    return None


def setup_libreoffice_macro():
    """Install the recalculation macro into LibreOffice's user directory."""
    system = platform.system()
    if system == "Darwin":
        macro_dir = os.path.expanduser(MACRO_DIR_MACOS)
    elif system == "Windows":
        macro_dir = MACRO_DIR_WINDOWS
    else:
        macro_dir = os.path.expanduser(MACRO_DIR_LINUX)

    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if os.path.exists(macro_file):
        try:
            if "RecalculateAndSave" in Path(macro_file).read_text():
                return True
        except Exception:
            pass

    soffice = get_soffice_path()
    if not soffice:
        return False

    if not os.path.exists(macro_dir):
        try:
            subprocess.run(
                [soffice, "--headless", "--terminate_after_init"],
                capture_output=True,
                timeout=15,
                env=get_soffice_env(),
            )
        except Exception:
            pass
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def scan_for_errors(filename):
    """Scan an Excel file for formula errors using openpyxl (no LibreOffice needed)."""
    try:
        wb = load_workbook(filename, data_only=True)
    except Exception as e:
        return {"error": f"Cannot open workbook: {e}"}

    excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
    error_details = {err: [] for err in excel_errors}
    total_errors = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    for err in excel_errors:
                        if err in cell.value:
                            location = f"{sheet_name}!{cell.coordinate}"
                            error_details[err].append(location)
                            total_errors += 1
                            break
    wb.close()

    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "error_summary": {},
    }

    for err_type, locations in error_details.items():
        if locations:
            result["error_summary"][err_type] = {
                "count": len(locations),
                "locations": locations[:20],
            }

    # Count formulas
    try:
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb_formulas.close()
        result["total_formulas"] = formula_count
    except Exception:
        result["total_formulas"] = -1

    return result


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    soffice = get_soffice_path()
    if not soffice:
        # No LibreOffice — scan for errors only, note formulas won't be recalculated
        result = scan_for_errors(abs_path)
        result["warning"] = (
            "LibreOffice not found — formulas were NOT recalculated. "
            "Open the file in Excel or LibreOffice to recalculate formulas, "
            "then re-run this script to verify."
        )
        result["recalculated"] = False
        return result

    if not setup_libreoffice_macro():
        result = scan_for_errors(abs_path)
        result["warning"] = (
            "Could not set up LibreOffice macro — formulas were NOT recalculated. "
            "Open the file in Excel or LibreOffice to recalculate."
        )
        result["recalculated"] = False
        return result

    cmd = [
        soffice,
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    timeout_cmd = has_timeout_cmd()
    if timeout_cmd:
        cmd = [timeout_cmd, str(timeout)] + cmd

    try:
        run_result = subprocess.run(
            cmd, capture_output=True, text=True,
            env=get_soffice_env(),
            timeout=timeout + 10 if not timeout_cmd else None,
        )
    except subprocess.TimeoutExpired:
        result = scan_for_errors(abs_path)
        result["warning"] = "LibreOffice timed out during recalculation."
        result["recalculated"] = False
        return result
    except Exception as e:
        result = scan_for_errors(abs_path)
        result["warning"] = f"LibreOffice error: {e}"
        result["recalculated"] = False
        return result

    if run_result.returncode != 0 and run_result.returncode != 124:
        error_msg = run_result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            result = scan_for_errors(abs_path)
            result["warning"] = "LibreOffice macro not configured properly. Formulas not recalculated."
            result["recalculated"] = False
            return result
        result = scan_for_errors(abs_path)
        result["warning"] = f"LibreOffice returned an error: {error_msg}"
        result["recalculated"] = False
        return result

    # Success — LibreOffice recalculated. Now scan for errors.
    result = scan_for_errors(abs_path)
    result["recalculated"] = True
    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("Falls back to error scanning if LibreOffice is unavailable")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - recalculated: true if LibreOffice ran, false if skipped")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
